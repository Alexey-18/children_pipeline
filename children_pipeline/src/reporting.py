from __future__ import annotations

import textwrap
from datetime import date
from pathlib import Path
from typing import Sequence

import matplotlib
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import pandas as pd

matplotlib.use("Agg")

from src.constants import (
    PLOTS_DIR,
    REPORT_CSV,
    REPORT_XLSX,
    RISK_HIGH,
    RISK_LOW,
    RISK_MEDIUM,
    SUMMARY_MD,
)

"""
Генерация отчетов: экспорт в CSV/Excel, графики динамики баллов и
краткая текстовая сводка в Markdown для руководителей.

Публичные функции
-----------------
* ``export_report(report_df, output_dir)``  — сохраняет CSV + Excel
* ``plot_dynamics(sessions_df, child_ids, output_dir)`` — один PNG на ребенка
* ``generate_summary(report_df, output_dir)``  — создает summary.md

Все функции идемпотентны: повторный запуск перезаписывает предыдущий вывод.
"""

# Внешний вид графиков
_RISK_COLORS: dict[str, str] = {
    RISK_HIGH: "#d62728",
    RISK_MEDIUM: "#ff7f0e", 
    RISK_LOW: "#1f77b4",
}
_DOMAIN_COLORS: dict[str, str] = {
    "Verbal_Request": "#2ca02c",
    "Listening": "#9467bd",
    "Social": "#8c564b",
    "Motor_Imitation": "#e377c2",
}
_SCORE_YMIN: int = 0
_SCORE_YMAX: int = 11

# Экспорт
def export_report(
    report_df: pd.DataFrame,
    output_dir: str | Path = "outputs",
) -> tuple[Path, Path]:
    """Экспортирует отчет о стагнации в CSV и Excel.
    Параметры
    --
    report_df:
        DataFrame, возвращаемый функцией :func:`src.analysis.detect_stagnation`.
    output_dir:
        Каталог для сохранения файлов. Создается при необходимости.

    Возвращает
    --
    tuple[Path, Path]
        Пути к созданным файлам CSV и Excel.
    """
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    csv_path = out / REPORT_CSV
    xlsx_path = out / REPORT_XLSX

    report_df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        report_df.to_excel(writer, sheet_name="stagnation_report", index=False)

        # Базовое условное форматирование: раскрашиваем колонку risk_level
        workbook = writer.book
        worksheet = writer.sheets["stagnation_report"]
        from openpyxl.styles import PatternFill, Font

        risk_fills = {
            RISK_HIGH: PatternFill("solid", fgColor="FFB3B3"),
            RISK_MEDIUM: PatternFill("solid", fgColor="FFE0A0"), 
            RISK_LOW: PatternFill("solid", fgColor="C8E6C9"),
        }

        risk_col_idx = (
            report_df.columns.tolist().index("risk_level") + 1
            if "risk_level" in report_df.columns
            else None
        )
        if risk_col_idx is not None:
            for row_idx, risk in enumerate(report_df["risk_level"], start=2):
                cell = worksheet.cell(row=row_idx, column=risk_col_idx)
                cell.fill = risk_fills.get(risk, PatternFill())
                cell.font = Font(bold=(risk == RISK_HIGH))

        # Автоподбор ширины колонок
        for col_cells in worksheet.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            worksheet.column_dimensions[col_cells[0].column_letter].width = (
                min(max_len + 2, 50)
            )

    print(f"[Экспорт] CSV  → {csv_path}")
    print(f"[Экспорт] XLSX → {xlsx_path}")
    return csv_path, xlsx_path

# Графики
def plot_dynamics(
    sessions_df: pd.DataFrame,
    child_ids: Sequence[str] | None = None,
    output_dir: str | Path = "outputs",
) -> list[Path]:
    """Строит графики динамики оценок для выбранных детей.
    Создается один PNG на ребенка, показывающий все области на одном графике.
    Параметры
    --
    sessions_df:
        Очищенный DataFrame сессий (результат валидатора).
    child_ids:
        Коды детей для построения графиков. По умолчанию все уникальные ``child_id``.
    output_dir:
        Корневой каталог вывода. Графики сохраняются в ``<output_dir>/plots/``.

    Возвращает
    --
    list[Path]
        Пути ко всем созданным PNG файлам.
    """
    plots_dir = Path(output_dir) / PLOTS_DIR
    plots_dir.mkdir(parents=True, exist_ok=True)

    if child_ids is None:
        child_ids = sessions_df["child_id"].dropna().unique().tolist()

    created: list[Path] = []

    for cid in child_ids:
        child_data = sessions_df[sessions_df["child_id"] == cid].copy()
        if child_data.empty:
            continue

        domains = child_data["domain"].unique()
        age = child_data["age"].iloc[0]
        diagnosis = child_data["diagnosis"].iloc[0]

        fig, ax = plt.subplots(figsize=(9, 4))

        for domain in sorted(domains):
            domain_data = (
                child_data[child_data["domain"] == domain]
                .sort_values("session_date")
                .drop_duplicates(subset=["session_date"])
            )
            color = _DOMAIN_COLORS.get(domain, "#333333")
            ax.plot(
                domain_data["session_date"],
                domain_data["assessment_score"],
                marker="o",
                linewidth=2,
                markersize=6,
                label=domain,
                color=color,
            )

        ax.set_title(
            f"{cid}  |  Возраст: {age}  |  {diagnosis}",
            fontsize=12,
            fontweight="bold",
        )
        ax.set_xlabel("Дата занятия", fontsize=10)
        ax.set_ylabel("Оценка навыка (1–10)", fontsize=10)
        ax.set_ylim(_SCORE_YMIN, _SCORE_YMAX)
        ax.yaxis.set_major_locator(mticker.MultipleLocator(1))
        ax.grid(axis="y", alpha=0.3, linestyle="--")
        ax.legend(loc="upper left", fontsize=9, framealpha=0.7)
        fig.autofmt_xdate(rotation=30)
        fig.tight_layout()

        png_path = plots_dir / f"{cid}.png"
        fig.savefig(png_path, dpi=120, bbox_inches="tight")
        plt.close(fig)
        created.append(png_path)

    print(f"[Графики] Сохранено {len(created)} PNG в {plots_dir}")
    return created

# Сводка
def generate_summary(
    report_df: pd.DataFrame,
    output_dir: str | Path = "outputs",
    min_days: int = 28,
) -> Path:
    """Создает Markdown-сводку для руководителей в ``summary.md``.
    Параметры
    --
    report_df:
        DataFrame, возвращаемый функцией :func:`src.analysis.detect_stagnation`.
    output_dir:
        Каталог для сохранения ``summary.md``.
    min_days:
        Значение ``min_days``, использованное при создании ``report_df``
        (только для отображения).
    Возвращает
    --
    Path
        Путь к созданному Markdown файлу.
    """
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    summary_path = out / SUMMARY_MD

    today = date.today().isoformat()
    total = len(report_df)
    n_high = (report_df["risk_level"] == RISK_HIGH).sum() if total else 0
    n_med = (report_df["risk_level"] == RISK_MEDIUM).sum() if total else 0
    n_low = (report_df["risk_level"] == RISK_LOW).sum() if total else 0

    lines: list[str] = [
        "# Отчёт об отсутствии прогресса",
        "",
        f"**Дата формирования:** {today}",
        f"**Окно анализа:** {min_days} дней",
        "",
        "## Сводка",
        "",
        "| Уровень риска | Кол-во случаев |",
        "|---------------|---------------|",
        f"| 🔴 HIGH       | {n_high}       |",
        f"| 🟡 MEDIUM     | {n_med}        |",
        f"| 🔵 LOW        | {n_low}        |",
        f"| **Итого**     | **{total}**    |",
        "",
    ]

    if total == 0:
        lines += [
            "## Результат",
            "",
            "> ✅ Случаев застоя не обнаружено за указанный период.",
            "",
        ]
    else:
        for risk, emoji in [(RISK_HIGH, "🔴"), (RISK_MEDIUM, "🟡"), (RISK_LOW, "🔵")]:
            subset = report_df[report_df["risk_level"] == risk]
            if subset.empty:
                continue
            lines += [
                f"## {emoji} {risk} — {len(subset)} {'случай' if len(subset)==1 else 'случаев'}",
                "",
            ]
            for _, row in subset.iterrows():
                comment = str(row.get("last_comment", "—"))
                if len(comment) > 120:
                    comment = comment[:117] + "..."
                stag_days = row.get("stagnation_days")
                stag_str = f"{stag_days} дн." if pd.notna(stag_days) else "н/д"

                lines += [
                    f"### {row['child_id']} — {row['domain']}",
                    "",
                    f"- **Возраст:** {row['age']}",
                    f"- **Диагноз:** {row['diagnosis']}",
                    f"- **Специалист:** {row.get('specialist_type', '—')}",
                    f"- **Дней без прогресса:** {stag_str}",
                    f"- **Оценка в начале окна → последняя:** "
                    f"{row['score_at_window_start']} → {row['score_latest']}",
                    f"- **Последняя заметка:** {comment}",
                    "",
                ]

    lines += [
        "###",
        "_Тестовое задание с СБЕР_",
    ]

    summary_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"[Сводка] Markdown → {summary_path}")
    return summary_path